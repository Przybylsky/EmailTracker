import datetime
import os

import openpyxl
import pandas as pd
import win32com.client as win32  # for outlook
from numpy import isnan


def DateToString(d=datetime.datetime.now(), sep=""):
    def getDigits(n):
        if n < 10:
            return "0" + str(n)
        else:
            return str(n)

    return getDigits(d.year) + sep + getDigits(d.month) + sep + getDigits(d.day)


def Row2string(row, header=False):
    _result = ''
    for i in range(0, len(row)):
        _result += '\t<th>' + str(row[i]) + '</th>\n' if header else '\t<td>' + str(row[i]) + '</td>\n'
    return _result


def SendEmail(to, subject, body, html):
    try:
        _mail = outlook.CreateItem(0)
        _mail.To = to
        _mail.cc = 'patryk.przybylski@nokia.com;bartosz.bieda@nokia.com;adrian.kuboszek@nokia.com'
        _mail.Subject = subject
        _mail.Body = body
        _mail.htmlBody = html
        _mail.Send()
        _mail = None
    except:
        return False

    print('Email sent to: ' + to)
    return True


def UpdateExcelFile(worksheet, pm):
    _rows = worksheet.max_row

    i = 2
    while i <= _rows:
        print(worksheet.cell(row=i, column=11).value)
        if (worksheet.cell(row=i, column=11).value == pm) & (worksheet.cell(row=i, column=1).value == 0):
            worksheet.cell(row=i, column=12).value = 'x'
        i += 1
    return

#colors
def get_html_from_df(df: object, columns: object, color: object) -> object:
    columns_order = df.columns.values
    result = pd.DataFrame()
    if not df.empty:
        # df["IPT_ID"] = df["IPT_ID"].astype(int).astype(str)
        df["Deal Id"] = df["Deal Id"].astype(int).astype(str)
    for index, row in df.iterrows():
        row_copy = row.copy()
        for column in columns:
            row_copy[column] = 'bgcolor' + str(row_copy[column])
        result = result.append(row_copy)
    if not result.empty:
        result = result[columns_order]
        return result.to_html(header=True, index=False).replace('>bgcolor', ' bgcolor=\"' + color + '\">')
    return ''


# setting up environment
root = 'C:\\Users\\pprzybyl\\Desktop\\IPT\\Reports\\Monthly\\06. June'
os.chdir(root)
fileName = 'email_tracker.xlsx'

# loading excel file to update
wb = openpyxl.load_workbook(fileName, data_only=True)  # root file
ws = wb.worksheets[0]  # worksheet with data

baseFile = pd.read_excel(fileName)
baseFile.head()
baseFile.shape

df = baseFile.loc[
    (baseFile['In_Hydra'] == 0) &
    (baseFile['Why?'] != 'Need reevaluation') &
    (baseFile['Copy_in_IPT'] == 0) &
    (baseFile['sent'].apply(lambda x: not x == x)) &
    (baseFile['Pricing Manager'] == baseFile['Pricing Manager'])].copy()

df['email'] = df['Pricing Manager'].apply(lambda x: str(x).replace(' ', '.') + '@nokia.com')
df['Opportunity ID Desc'] = df['Opportunity ID Desc'].apply(lambda x: '' if x != x else x[1:])

outlook = win32.Dispatch('outlook.application')  # start outlook application

columns = ['Deal Id', 'Total Revenue', 'Customer name', 'Workteam', 'Opportunity ID Desc', 'Updated At', 'Pricing Manager']

# df = df[columns]
invalid_customer = pd.DataFrame()
invalid_opp_ID = pd.DataFrame()
invalid_both = pd.DataFrame()
receivers = []

for index, row in df.iterrows():
    if row['Pricing Manager'] in receivers:
        pass
    else:
        receivers.append(row['Pricing Manager'])

    if "Dummy Customer" in row["Customer name"]:
        if "Dummy Customer" in row["Customer name"] and "wrong" in row["Opportunity ID Desc"]:
            invalid_both = invalid_both.append(row)
        elif "Dummy Customer" in row["Customer name"]:
            invalid_customer = invalid_customer.append(row)
    #elif "Dummy Customer" in row["Customer name"] and "" in row["Opportunity ID Desc"]:
        #both = both.append(row)
    elif "" in row["Opportunity ID Desc"]:
        invalid_opp_ID = invalid_opp_ID.append(row)

invalid_opp_ID = invalid_opp_ID[columns]
invalid_customer = invalid_customer[columns]
invalid_both = invalid_both[columns]

path = r'C:\Users\pprzybyl\Desktop\IPT\Reports\Monthly\06. June'
writer = pd.ExcelWriter('Email_Backlog.xlsx')

for receiver in receivers:
    print(receiver + '\n')
    invalid_customer_sent = invalid_customer[invalid_customer["Pricing Manager"] == receiver]
    invalid_opp_ID_sent = invalid_opp_ID[invalid_opp_ID['Pricing Manager'] == receiver]
    invalid_both_sent = invalid_both[invalid_both['Pricing Manager'] == receiver]

    _header = '<tr>\n' + Row2string(columns, True) + '\n</tr>\n'  # generate table head
    _text = "Hi, <br><br>  The following cases, which were uploaded to IPT didn't get to Hydra due to lack of Opportunity ID, Customer ID (Dummy Customer) or wrong Revenue.<br><br>\
             There is a necessity to investigate that so please provide me with information on these cases below: <br>"
    if not invalid_customer_sent.empty:
        print(invalid_customer.head(n=1))
        _text += "<br> <b>- Invalid Customer Name has been detected in the cases with the following Deal IDs and highlighted below:</b>"
        _text += get_html_from_df(invalid_customer_sent, ['Customer name'], 'orange')
    if not invalid_opp_ID_sent.empty:
        _text += "<br> <b>- Invalid or lack of Opportunity ID has been detected in the cases with the following Deal IDs and highlighted below:</b> <br>"
        _text += get_html_from_df(invalid_opp_ID_sent, ['Opportunity ID Desc'], 'orange')
    if not invalid_both_sent.empty:
        _text += "<br> <b>- Invalid or lack Opportunity ID and Customer Name has been detected in the cases with the following Deal IDs and highlighted below: </b> <br>"
        _text += get_html_from_df(invalid_both_sent, ['Customer name','Opportunity ID Desc'], 'orange')
    _text += "<br><br> Best regards, <br> Patryk Przybylski"

    # for index, row in data[columns].iterrows():
    _emailSent = SendEmail(to=str(receiver), subject='Invalid IPT Cases - 06/2018', body='', html=_text)

    if _emailSent:
        baseFile.loc[(baseFile['Pricing Manager'] == receiver), "sent"] = 'x'

    invalid_customer_sent = None
    invalid_opp_ID_sent = None
    invalid_both_sent = None

baseFile.to_excel(writer, index=False)
writer.save()
# wb.save('email_tracker_' + DateToString() + '.xlsx')
outlook = None
# wb = None
